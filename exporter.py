"""
Excel Exporter — writes scraped LinkedIn profile data to .xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(profiles: list[dict], output_path: str) -> None:
    """
    Export a list of profile dicts to a styled Excel workbook.

    Each dict should have keys:
        name, email, phone, profile_link, college_name, course
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn Profiles"

    # ── Column headers ──────────────────────────────
    headers = ["S.No", "Name", "Email", "Phone", "Profile Link", "College Name", "Course"]
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Data rows ───────────────────────────────────
    data_font = Font(name="Calibri", size=11)
    link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")
    data_alignment = Alignment(vertical="center", wrap_text=True)

    for row_idx, profile in enumerate(profiles, start=2):
        serial = row_idx - 1
        row_data = [
            serial,
            profile.get("name", ""),
            profile.get("email", ""),
            profile.get("phone", ""),
            profile.get("profile_link", ""),
            profile.get("college_name", ""),
            profile.get("course", ""),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = data_alignment

            # Make profile link clickable
            if col_idx == 5 and value:
                cell.font = link_font
                cell.hyperlink = value
            else:
                cell.font = data_font

    # ── Auto-fit column widths ──────────────────────
    min_widths = [6, 25, 30, 18, 50, 35, 25]
    for col_idx, min_w in enumerate(min_widths, start=1):
        max_len = min_w
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, 60)

    # ── Freeze header row ───────────────────────────
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n✅  Saved {len(profiles)} profiles to: {output_path}")
